"""
Fleet module views.

Provides views for displaying fleet module cards, KPIs,
detailed module information with maintenance projections,
and sync controls for the Access-to-Postgres ETL.
"""

import logging
import subprocess
import sys

from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET, require_POST

from core.services.grid_projection import (
    CYCLE_COLORS,
    DEFAULT_AVG_MONTHLY_KM,
    DEFAULT_MONTHS,
    GridProjectionService,
)
from core.services.maintenance_projection import (
    CSR_HEAVY_CYCLES,
    MaintenanceProjectionService,
    TOSHIBA_HEAVY_CYCLES,
)
from etl.extractors.access_extractor import (
    get_module_detail_from_access,
    get_modules_with_fallback,
)
from etl.extractors.postgres_extractor import (
    get_module_detail_from_postgres,
    is_postgres_staging_available,
)

from .stub_data import get_fleet_summary

logger = logging.getLogger(__name__)


@require_GET
def module_list(request):
    """
    Display all fleet modules as cards with summary KPIs.

    Data source priority:
    1. Access database (if configured and available)
    2. Stub data (fallback for development/CI)

    URL: /fleet/modules/
    """
    # Get modules from Access or fallback to stub
    modules = get_modules_with_fallback()
    summary = get_fleet_summary(modules)

    # Optional filtering by fleet type
    fleet_filter = request.GET.get("fleet", "all")
    if fleet_filter == "csr":
        modules = [m for m in modules if m.fleet_type == "CSR"]
    elif fleet_filter == "toshiba":
        modules = [m for m in modules if m.fleet_type == "Toshiba"]

    # Sync status for the sync bar
    from infrastructure.database.models import SyncLog
    latest_sync = SyncLog.objects.first()  # ordered by -started_at

    context = {
        "modules": modules,
        "summary": summary,
        "fleet_filter": fleet_filter,
        "latest_sync": latest_sync,
    }

    return render(request, "fleet/module_list.html", context)


@require_GET
def module_detail(request, module_id: str):
    """
    Display detailed information for a single maintenance module.

    Shows:
    - Module card info (same as list view)
    - Projection of next intervention
    - Key data per cycle type (datos clave para proyección)
    - Maintenance history (last 365 days)

    URL: /fleet/modules/<module_id>/

    When data comes from Access (module_db_id is set), history and
    key_data are loaded on demand from the legacy database.
    """
    # Normalize module_id to uppercase
    module_id = module_id.upper()

    # Get all modules (same data source as list view)
    all_modules = get_modules_with_fallback()

    # Build module lookup and list for dropdown
    module_lookup = {m.module_id: m for m in all_modules}
    module = module_lookup.get(module_id)

    if module is None:
        raise Http404(f"Módulo {module_id} no encontrado")

    # Load detail data on demand if not already populated
    # (Access-sourced modules have module_db_id but empty history/key_data)
    if not module.maintenance_key_data and module.module_db_id is not None:
        # Prefer Postgres staging; fall back to live ODBC
        if is_postgres_staging_available():
            logger.info(
                "Loading detail data from Postgres for %s (db_id=%s)",
                module_id, module.module_db_id,
            )
            history, key_data = get_module_detail_from_postgres(
                module_db_id=module.module_db_id,
                module_id=module.module_id,
                fleet_type=module.fleet_type,
                km_total=module.km_total_accumulated,
                commissioning_date=module.reference_date,
            )
        else:
            logger.info(
                "Loading detail data from Access for %s (db_id=%s)",
                module_id, module.module_db_id,
            )
            history, key_data = get_module_detail_from_access(
                module_db_id=module.module_db_id,
                module_id=module.module_id,
                fleet_type=module.fleet_type,
                km_total=module.km_total_accumulated,
            )
        module.maintenance_history = history
        module.maintenance_key_data = key_data

    # Prepare key data as dicts for the projection service
    key_data_dicts = [
        {
            "cycle_type": kd.cycle_type,
            "cycle_km": kd.cycle_km,
            "km_at_last": kd.km_at_last,
            "last_date": kd.last_date,
        }
        for kd in module.maintenance_key_data
    ]

    # Calculate next intervention projection
    projection = MaintenanceProjectionService.project_next_intervention(
        fleet_type=module.fleet_type,
        km_total=module.km_total_accumulated,
        key_data=key_data_dicts,
    )

    # Module list for the dropdown selector
    module_options = [
        {"id": m.module_id, "label": f"{m.module_id} ({m.fleet_type})"}
        for m in all_modules
    ]

    # Determine data source for display
    if module.module_db_id is not None:
        data_source = "POSTGRES" if is_postgres_staging_available() else "ACCESS"
    else:
        data_source = "STUB"

    context = {
        "module": module,
        "projection": projection,
        "module_options": module_options,
        "current_module_id": module_id,
        "data_source": data_source,
    }

    return render(request, "fleet/module_detail.html", context)


# ---------------------------------------------------------------------------
# Sync Access → Postgres
# ---------------------------------------------------------------------------

@require_GET
def sync_status(request):
    """
    Return JSON with the latest sync status for Alpine.js polling.

    URL: /fleet/sync/status/
    """
    from infrastructure.database.models import SyncLog

    latest = SyncLog.objects.first()  # ordered by -started_at
    if latest is None:
        return JsonResponse({
            "has_synced": False,
            "status": "never",
            "started_at": None,
            "finished_at": None,
            "duration": None,
            "tables": {},
            "error": "",
        })

    return JsonResponse({
        "has_synced": True,
        "status": latest.status,
        "started_at": latest.started_at.isoformat() if latest.started_at else None,
        "finished_at": latest.finished_at.isoformat() if latest.finished_at else None,
        "duration": latest.duration_seconds,
        "tables": latest.tables_synced or {},
        "error": latest.error_message or "",
    })


@require_POST
def sync_trigger(request):
    """
    Launch ``sync_access`` as a detached subprocess and return immediately.

    The subprocess writes its progress into a ``SyncLog`` row that the
    frontend polls via ``/fleet/sync/status/``.

    URL: POST /fleet/sync/trigger/
    """
    from infrastructure.database.models import SyncLog

    # Prevent concurrent syncs
    running = SyncLog.objects.filter(status="running").exists()
    if running:
        return JsonResponse(
            {"ok": False, "error": "Ya hay una sincronización en curso."},
            status=409,
        )

    # Launch sync_access as a detached subprocess
    python_exe = sys.executable  # same interpreter running Django
    cmd = [python_exe, "manage.py", "sync_access"]

    logger.info("Launching sync_access subprocess: %s", " ".join(cmd))

    try:
        subprocess.Popen(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            # On Windows, CREATE_NO_WINDOW prevents a console flash
            creationflags=(
                subprocess.CREATE_NO_WINDOW
                if sys.platform == "win32"
                else 0
            ),
        )
    except Exception as e:
        logger.exception("Failed to launch sync_access subprocess")
        return JsonResponse(
            {"ok": False, "error": f"Error al lanzar sync: {e}"},
            status=500,
        )

    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Maintenance Planner (projection grid)
# ---------------------------------------------------------------------------

@require_GET
def projection_grid(request):
    """
    Display the maintenance projection grid (Excel-like).

    Shows accumulated km per heavy cycle per module, projected month
    by month, with colour-coded cells when the threshold is exceeded.

    Query parameters:
        fleet: "csr" or "toshiba" (required)
        months: number of months to project (default 18)
        avg_km: average monthly km override (optional)

    URL: /fleet/planner/
    """
    from datetime import date as _date

    fleet_param = request.GET.get("fleet", "csr").lower()
    fleet_type = "CSR" if fleet_param == "csr" else "Toshiba"

    months = int(request.GET.get("months", DEFAULT_MONTHS))
    months = max(1, min(months, 60))  # clamp 1-60

    default_avg = DEFAULT_AVG_MONTHLY_KM[fleet_type]
    avg_km = int(request.GET.get("avg_km", default_avg))

    # Get modules and filter by fleet
    all_modules = get_modules_with_fallback()
    fleet_modules = [m for m in all_modules if m.fleet_type == fleet_type]

    today = _date.today()

    # Build modules_data for GridProjectionService
    modules_data: list[dict] = []
    for mod in fleet_modules:
        # Load key_data if not populated (Access-sourced modules)
        if not mod.maintenance_key_data and mod.module_db_id is not None:
            if is_postgres_staging_available():
                _, key_data = get_module_detail_from_postgres(
                    module_db_id=mod.module_db_id,
                    module_id=mod.module_id,
                    fleet_type=mod.fleet_type,
                    km_total=mod.km_total_accumulated,
                    commissioning_date=mod.reference_date,
                )
                mod.maintenance_key_data = key_data
            else:
                _, key_data = get_module_detail_from_access(
                    module_db_id=mod.module_db_id,
                    module_id=mod.module_id,
                    fleet_type=mod.fleet_type,
                    km_total=mod.km_total_accumulated,
                )
                mod.maintenance_key_data = key_data

        # Convert MaintenanceKeyData to dicts for the grid service
        kd_list = [
            {
                "cycle_type": kd.cycle_type,
                "cycle_km": kd.cycle_km,
                "km_since": kd.km_since,
            }
            for kd in mod.maintenance_key_data
        ]

        modules_data.append({
            "module_id": mod.module_id,
            "fleet_type": mod.fleet_type,
            "key_data": kd_list,
        })

    # Generate grid
    grid = GridProjectionService.generate_grid(
        modules_data=modules_data,
        avg_monthly_km=avg_km,
        months=months,
        reference_date=today,
    )
    month_headers = GridProjectionService.get_month_headers(
        months=months,
        reference_date=today,
    )

    context = {
        "grid": grid,
        "month_headers": month_headers,
        "fleet_type": fleet_type,
        "fleet_param": fleet_param,
        "months": months,
        "avg_km": avg_km,
        "avg_km_csr": DEFAULT_AVG_MONTHLY_KM["CSR"],
        "avg_km_toshiba": DEFAULT_AVG_MONTHLY_KM["Toshiba"],
    }

    return render(request, "fleet/projection_grid.html", context)


@require_GET
def projection_export(request):
    """
    Export the projection grid to an Excel file with colour formatting.

    Reuses the same grid generation logic as ``projection_grid`` and
    writes it to an openpyxl workbook.

    Query parameters: same as ``projection_grid``.

    URL: /fleet/planner/export/
    """
    from datetime import date as _date
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fleet_param = request.GET.get("fleet", "csr").lower()
    fleet_type = "CSR" if fleet_param == "csr" else "Toshiba"
    months = int(request.GET.get("months", DEFAULT_MONTHS))
    months = max(1, min(months, 60))
    default_avg = DEFAULT_AVG_MONTHLY_KM[fleet_type]
    avg_km = int(request.GET.get("avg_km", default_avg))

    all_modules = get_modules_with_fallback()
    fleet_modules = [m for m in all_modules if m.fleet_type == fleet_type]
    today = _date.today()

    # Build modules_data (same as grid view)
    modules_data: list[dict] = []
    for mod in fleet_modules:
        if not mod.maintenance_key_data and mod.module_db_id is not None:
            if is_postgres_staging_available():
                _, key_data = get_module_detail_from_postgres(
                    module_db_id=mod.module_db_id,
                    module_id=mod.module_id,
                    fleet_type=mod.fleet_type,
                    km_total=mod.km_total_accumulated,
                    commissioning_date=mod.reference_date,
                )
                mod.maintenance_key_data = key_data
            else:
                _, key_data = get_module_detail_from_access(
                    module_db_id=mod.module_db_id,
                    module_id=mod.module_id,
                    fleet_type=mod.fleet_type,
                    km_total=mod.km_total_accumulated,
                )
                mod.maintenance_key_data = key_data

        kd_list = [
            {
                "cycle_type": kd.cycle_type,
                "cycle_km": kd.cycle_km,
                "km_since": kd.km_since,
            }
            for kd in mod.maintenance_key_data
        ]
        modules_data.append({
            "module_id": mod.module_id,
            "fleet_type": mod.fleet_type,
            "key_data": kd_list,
        })

    grid = GridProjectionService.generate_grid(
        modules_data=modules_data,
        avg_monthly_km=avg_km,
        months=months,
        reference_date=today,
    )
    month_headers = GridProjectionService.get_month_headers(
        months=months, reference_date=today,
    )

    # --- Build Excel workbook ---
    # Colour mapping: Tailwind class -> openpyxl hex (no leading #)
    FILL_MAP = {
        "AN": PatternFill("solid", fgColor="DCFCE7"),  # green-100
        "BA": PatternFill("solid", fgColor="FEF9C3"),  # yellow-100
        "PE": PatternFill("solid", fgColor="E0F2FE"),  # sky-100
        "DA": PatternFill("solid", fgColor="FEE2E2"),  # red-100
        "RB": PatternFill("solid", fgColor="FEF9C3"),  # yellow-100
        "RG": PatternFill("solid", fgColor="FEE2E2"),  # red-100
    }
    FONT_MAP = {
        "AN": Font(color="166534", bold=True),   # green-800
        "BA": Font(color="854D0E", bold=True),   # yellow-800
        "PE": Font(color="075985", bold=True),    # sky-800
        "DA": Font(color="991B1B", bold=True),    # red-800
        "RB": Font(color="854D0E", bold=True),   # yellow-800
        "RG": Font(color="991B1B", bold=True),    # red-800
    }

    wb = Workbook()
    ws = wb.active
    ws.title = f"Proyeccion {fleet_type}"

    # Header row
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="F3F4F6")  # gray-100
    headers = ["Modulo", "Ciclo", "Umbral"] + month_headers

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    current_row = 2
    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")

    for module in grid:
        for row_idx, cycle_row in enumerate(module.cycle_rows):
            # Module ID (first cycle only)
            if row_idx == 0:
                mod_cell = ws.cell(
                    row=current_row, column=1,
                    value=module.module_id,
                )
                mod_cell.font = Font(bold=True)
                if len(module.cycle_rows) > 1:
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=1,
                        end_row=current_row + len(module.cycle_rows) - 1,
                        end_column=1,
                    )
                    mod_cell.alignment = Alignment(
                        horizontal="center", vertical="center",
                    )

            # Cycle type
            ws.cell(
                row=current_row, column=2,
                value=cycle_row.cycle_type,
            ).alignment = center_align

            # Threshold
            ws.cell(
                row=current_row, column=3,
                value=cycle_row.cycle_km,
            ).alignment = right_align
            ws.cell(row=current_row, column=3).number_format = "#,##0"

            # Month cells
            for m_idx, mp in enumerate(cycle_row.months):
                cell = ws.cell(
                    row=current_row, column=4 + m_idx,
                    value=mp.km_accumulated,
                )
                cell.alignment = right_align
                cell.number_format = "#,##0"

                if mp.exceeded:
                    fill = FILL_MAP.get(cycle_row.cycle_type)
                    font = FONT_MAP.get(cycle_row.cycle_type)
                    if fill:
                        cell.fill = fill
                    if font:
                        cell.font = font

            current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    for col in range(4, 4 + months):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Freeze panes (header + module/cycle/threshold columns)
    ws.freeze_panes = "D2"

    # Write to response
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"proyeccion_{fleet_type.lower()}_{today.isoformat()}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
